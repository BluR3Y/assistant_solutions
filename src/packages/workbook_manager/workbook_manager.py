import os
import pandas as pd
import openpyxl
import pathlib
import warnings
import numpy as np

from typing import Union

from . import PropertyManager, WorkbookLogManager

class WorkbookManager:
    read_file_path = None
    log_manager = None
    df = {}
    
    def __init__(self, read_file_path: str = None, create_sheets: dict = None):
        if read_file_path:
            if not os.path.exists(read_file_path):
                raise ValueError(f"The WorkbookManager was provided an invalid file path: {read_file_path}")
            
            # Store the file path
            self.read_file_path = read_file_path
            
            # Read the contents of the workbook
            self.df = pd.read_excel(read_file_path, sheet_name=None)
            
            file_path_obj = pathlib.Path(read_file_path)
            log_file_path = os.path.join(file_path_obj.parent, f"{file_path_obj.stem}_workbook_logs.json")
            self.log_manager = WorkbookLogManager(log_file_path)
        elif create_sheets:
            self.df = {sheet_name: pd.DataFrame(sheet_rows) for sheet_name, sheet_rows in create_sheets.items()}
        
        # Initialize an instance of the Comment Manager
        self.property_manager = PropertyManager(read_file_path, self.df.keys())
        
    def create_sheet(self, sheet_name: str, sheet_columns: list):
        if sheet_name in self.df.keys():
            raise ValueError(f"A sheet with the name '{sheet_name}' already exists in the workbook.")
        
        sheet_data_frame = None
        populated_data = all(isinstance(item, dict) for item in sheet_columns)
        if populated_data:
            sheet_data_frame = pd.DataFrame(sheet_columns)
        else:
            sheet_data_frame = pd.DataFrame(columns=sheet_columns)
            
        self.df[sheet_name] = sheet_data_frame
        self.property_manager.property_store[sheet_name] = {}
        
    def update_cell(self, process: str, sheet: str, row: int, col: int, new_val):
        if sheet not in self.df:
            raise ValueError(f"The sheet with the name '{sheet}' does not exist in the workbook.")
        
        sheet_data_frame = self.df[sheet]
        num_rows, num_cols = sheet_data_frame.shape
        # Ensure row and col are within valid bounds
        if not (0 <= row < num_rows) or not (0 <= col < num_cols):
            raise IndexError(f"Row {row} or Column {col} is out of bound for the sheet '{sheet}'.")
        
        cell_prev_value = sheet_data_frame.iat[row, col]
        sheet_data_frame.iat[row, col] = new_val

        if self.log_manager:
            self.log_manager.append_log(
                process,
                sheet,
                row,
                col,
                cell_prev_value,
                new_val
            )
        
    def append_row(self, sheet: str, props: dict):
        # Create a new DataFrame
        new_row = pd.DataFrame({key: [value] for key, value in props.items()})
        # Append using pd.concat
        self.df[sheet] = pd.concat([self.df[sheet], new_row], ignore_index=True)
        
    def row_follows_condition(self, row, conditions: dict) -> bool:
        for identifier, value in conditions.items():
            if row[identifier] != value:
                return False
        return True
        
    # def get_entries(self, sheet_name: str, conditions: dict, all: bool = False):
    #     """Retrieve rows from the given sheet matching conditions."""
    #     if sheet_name not in self.df:
    #         raise ValueError(f"The sheet '{sheet_name}' was not found in the workbook.")
        
    #     try:
    #         sheet_data_frame = self.df[sheet_name]
            
    #         # Apply filtering directly
    #         mask = pd.Series(True, index=sheet_data_frame.index)
    #         for column, value in conditions.items():
    #             mask &= (sheet_data_frame[column] == value)
                
    #         filtered_rows = sheet_data_frame[mask]
            
    #         if filtered_rows.empty:
    #             return None
            
    #         filtered_rows.replace([pd.NaT, np.nan], None, inplace=False)
            
    #         return filtered_rows.to_dict() if all else filtered_rows.iloc[0].to_dict()
    def get_entries(self, sheet_name: str, conditions: dict, return_all: bool = False):
        """Retreive rows from the given sheet matching conditions."""
        
        if sheet_name not in self.df:
            raise ValueError(f"The sheet '{sheet_name}' was not found in the workbook.")
        
        sheet_data_frame = self.df[sheet_name]
        
        # Ensure that conditions refer to existing columns
        invalid_columns = [col for col in conditions if col not in sheet_data_frame.columns]
        if invalid_columns:
            raise KeyError(f"Invalid column names in conditions: {invalid_columns}")
        
        # Apply filtering
        mask = pd.Series(True, index=sheet_data_frame.index)
        for column, value in conditions.items():
            mask &= (sheet_data_frame[column] == value)
        
        filtered_rows = sheet_data_frame[mask]
        
        if filtered_rows.empty:
            return None
        
        # Replace NaN and NaT with None
        filtered_rows = filtered_rows.replace([pd.NaT, np.nan], None)
        
        return filtered_rows.to_dict(orient="records") if return_all else filtered_rows.iloc[0].to_dict()
    
    def save_changes(self, write_file_path: str, index=False):
        """
        Saves the data stored in the pandas dataframes by converting each dataframe into an excel sheet in the same workbook.
        Additionally, function also sets the 'sheet_state' property to that of the sheet in the workbook that was imported.
        
        # Parameters:
        - write_file_path: file path where the migration data will be stored.
        - index: Will determine if the index column from the DataFrame will persist in the stored excel sheet.
        """
                
        if write_file_path == self.read_file_path:
            warnings.warn(f"The save path is the same as the read path, meaning that file will be overwritten with changes.")
        
        # populated_sheets = {sheet_name: sheet_content for sheet_name, sheet_content in self.df.items() if not sheet_content.empty}
        is_empty = all([sheet_content.empty for sheet_content in self.df.values()])
        if is_empty:
            print("Workbook is empty. Now exiting Workbook Manager.")
        
        try:
            # Use ExcelWriter to write multiple sheets back into the Excel file
            with pd.ExcelWriter(write_file_path, engine='openpyxl', mode='w') as writer:
                for sheet_name, df_sheet in self.df.items():
                    df_sheet.to_excel(writer, sheet_name=sheet_name, index=index)
                    
            # Save LogManager changes
            if self.log_manager:
                self.log_manager.save_logs()
                
            # Save Workbook Properties
            self.property_manager.apply_changes(write_file_path)
        except Exception as err:
            raise Exception(f"Error occured while attempting to save Workbook data: {err}")